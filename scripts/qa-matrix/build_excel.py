"""Build the QA test-matrix workbook from qa-apis.json + metadata.json.
Output: tests/QA-Test-Matrix.xlsx  (sheets: QA Test Matrix, Epics & Features)."""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
HERE = os.path.dirname(__file__)
BASE_URL = "https://projexlight.com/s/"

apis = json.load(open(os.path.join(HERE, "qa-apis.json"), encoding="utf-8"))
meta = json.load(open(os.path.join(HERE, "metadata.json"), encoding="utf-8"))
features, scenarios = meta["features"], meta.get("scenarios", {})

# Merge the full 34-epic catalogue so every referenced epic resolves a title.
epics = {}
_all = os.path.join(HERE, "all-epics.json")
if os.path.exists(_all):
    _raw = json.load(open(_all, encoding="utf-8"))
    _items = _raw.get("data", {}).get("items", _raw) if isinstance(_raw, dict) else _raw
    for e in _items:
        epics[e["id"]] = {
            "short_id": e.get("short_id", ""), "title": e.get("title", ""),
            "source_module": (e.get("customdata") or {}).get("source_module"),
        }
epics.update(meta["epics"])  # api-referenced epics (already same shape)


def url(short):
    return f"{BASE_URL}{short}" if short else ""


def jdump(v):
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, indent=2, ensure_ascii=False)
    return str(v)


def scen_text(feature_id):
    rows = scenarios.get(feature_id, []) or []
    names = " | ".join(s.get("title", "") for s in rows)
    urls = " | ".join(url(s.get("short_id", "")) for s in rows if s.get("short_id"))
    return names, urls


# ---- styling ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BAND_FILLS = [PatternFill("solid", fgColor="EAF1FB"), PatternFill("solid", fgColor="FFFFFF")]
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")

wb = Workbook()

# ============================ Sheet 1: QA Test Matrix ============================
ws = wb.active
ws.title = "QA Test Matrix"
COLS = [
    ("Epic", 26), ("Epic URL", 30), ("Feature", 30), ("Feature URL", 30),
    ("Scenario(s)", 30), ("Scenario URL(s)", 30), ("SDK / Service", 22),
    ("API URI", 34), ("Method", 8), ("Auth", 6), ("Test Case", 26),
    ("Payload", 40), ("Path Params", 22), ("Exp. Status", 10),
    ("Expected Result", 40), ("Depends On", 28), ("Testability", 24), ("Source File", 38),
]
ws.append([c[0] for c in COLS])
for i, (_name, w) in enumerate(COLS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
    cell = ws.cell(row=1, column=i)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

# Sort: epic short_id, feature title, endpoint, method
def sort_key(r):
    e = epics.get(r["epicId"], {})
    f = features.get(r["featureId"], {})
    return (e.get("short_id", "ZZ"), f.get("title", "zz"), r["endpoint"], r["method"])

apis_sorted = sorted(apis, key=sort_key)

# Colour-band per epic for readability.
last_epic = None
band = 0
for r in apis_sorted:
    e = epics.get(r["epicId"], {})
    f = features.get(r["featureId"], {})
    if e.get("short_id") != last_epic:
        band ^= 1
        last_epic = e.get("short_id")
    sc_names, sc_urls = scen_text(r["featureId"])
    auth = "Yes" if r.get("requiresAuth") else ("No" if r.get("requiresAuth") is False else "")
    testability = r.get("testability", "")
    if r.get("skipReason"):
        testability = (testability + " — " + r["skipReason"]).strip(" —")
    row = [
        e.get("title", "(unmapped epic)"), url(e.get("short_id", "")),
        f.get("title", "(unmapped feature)"), url(f.get("short_id", "")),
        sc_names, sc_urls, r.get("sdk", ""),
        r["endpoint"], r["method"], auth, r.get("case", ""),
        jdump(r.get("payload")), jdump(r.get("pathParams")), str(r.get("expectedStatus", "")),
        jdump(r.get("expectedResponse")), jdump(r.get("dependsOn")), testability, r.get("file", ""),
    ]
    ws.append(row)
    ridx = ws.max_row
    for cidx in range(1, len(COLS) + 1):
        cell = ws.cell(row=ridx, column=cidx)
        cell.fill = BAND_FILLS[band]
        cell.border = BORDER
        # wrap the json-ish / long columns
        cell.alignment = WRAP_TOP if cidx in (1, 3, 5, 6, 8, 11, 12, 13, 15, 16, 17, 18) else TOP
    # hyperlinks on URL columns
    for cidx, val in ((2, row[1]), (4, row[3]), (6, row[5])):
        if val and " | " not in val:
            ws.cell(row=ridx, column=cidx).hyperlink = val
            ws.cell(row=ridx, column=cidx).font = Font(color="0563C1", underline="single")

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"

# ============================ Sheet 2: Epics & Features index ============================
ws2 = wb.create_sheet("Epics & Features")
ICOLS = [("Epic", 30), ("Epic URL", 30), ("SDK / source", 30), ("Feature", 38),
         ("Feature URL", 30), ("# API cases", 12), ("Scenario(s)", 40)]
ws2.append([c[0] for c in ICOLS])
for i, (_n, w) in enumerate(ICOLS, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
    cell = ws2.cell(row=1, column=i)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

# count api cases per feature
cnt = {}
for r in apis:
    cnt[r["featureId"]] = cnt.get(r["featureId"], 0) + 1

# group features by epic
by_epic = {}
for fid, f in features.items():
    by_epic.setdefault(f.get("epic_id", ""), []).append((fid, f))

band = 0
for eid in sorted(by_epic, key=lambda x: epics.get(x, {}).get("short_id", "ZZ")):
    e = epics.get(eid, {})
    band ^= 1
    for fid, f in sorted(by_epic[eid], key=lambda x: x[1].get("title", "")):
        sc_names, _ = scen_text(fid)
        ws2.append([
            e.get("title", ""), url(e.get("short_id", "")), e.get("source_module") or "",
            f.get("title", ""), url(f.get("short_id", "")), cnt.get(fid, 0), sc_names,
        ])
        ridx = ws2.max_row
        for cidx in range(1, len(ICOLS) + 1):
            c = ws2.cell(row=ridx, column=cidx)
            c.fill = BAND_FILLS[band]
            c.border = BORDER
            c.alignment = WRAP_TOP if cidx in (1, 4, 7) else TOP
        for cidx, val in ((2, ws2.cell(row=ridx, column=2).value), (5, ws2.cell(row=ridx, column=5).value)):
            if val:
                ws2.cell(row=ridx, column=cidx).hyperlink = val
                ws2.cell(row=ridx, column=cidx).font = Font(color="0563C1", underline="single")

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(ICOLS))}{ws2.max_row}"

out = os.path.join(ROOT, "tests", "QA-Test-Matrix.xlsx")
try:
    wb.save(out)
except PermissionError:
    out = os.path.join(ROOT, "tests", "QA-Test-Matrix-latest.xlsx")
    wb.save(out)
    print("NOTE: QA-Test-Matrix.xlsx was locked (open in Excel?) — wrote to QA-Test-Matrix-latest.xlsx instead.")
print(f"wrote {out}  |  matrix rows={len(apis_sorted)}  features={len(features)}  epics={len(epics)}")
